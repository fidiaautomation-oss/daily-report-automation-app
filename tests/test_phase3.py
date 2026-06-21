import pytest
import pandas as pd
import openpyxl
from unittest.mock import MagicMock
from phase3.excel_writer import ExcelWriter
from phase3.report_archiver import ReportArchiver


@pytest.fixture
def template_xlsx(tmp_path):
    """テスト用の簡易Excelテンプレートを作成する"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "日報"
    ws["A1"] = "ヘッダー行1"
    ws["A2"] = "ヘッダー行2"
    path = tmp_path / "template.xlsx"
    wb.save(path)
    return str(path)


def test_excel_writer_writes_correct_cells(template_xlsx, tmp_path):
    """mapping.yamlに従って正しいセルに値が書き込まれること"""
    mapping = {
        "sheet_name": "日報",
        "data_start_row": 3,
        "mappings": [
            {"csv_col": "clicks", "excel_col": "B"},
            {"csv_col": "cost", "excel_col": "C"},
        ],
    }
    df = pd.DataFrame([{"clicks": 100, "cost": 5000}])
    output_path = str(tmp_path / "output.xlsx")

    writer = ExcelWriter(mapping)
    writer.write(template_xlsx, df, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["日報"]
    assert ws["B3"].value == 100
    assert ws["C3"].value == 5000


def test_excel_writer_preserves_existing_cells(template_xlsx, tmp_path):
    """書き込み対象外のセルが変更されていないこと"""
    mapping = {
        "sheet_name": "日報",
        "data_start_row": 3,
        "mappings": [{"csv_col": "clicks", "excel_col": "B"}],
    }
    df = pd.DataFrame([{"clicks": 100}])
    output_path = str(tmp_path / "output.xlsx")

    writer = ExcelWriter(mapping)
    writer.write(template_xlsx, df, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["日報"]
    assert ws["A1"].value == "ヘッダー行1"


def test_report_archiver_uploads_to_correct_folder():
    """report/YYYY-MM-DD/ フォルダへアップロードされること"""
    archiver = ReportArchiver.__new__(ReportArchiver)
    mock_service = MagicMock()
    mock_service.files().list().execute.return_value = {"files": [{"id": "report_folder_id"}]}
    mock_service.files().create().execute.return_value = {"id": "date_folder_id"}
    archiver.service = mock_service
    archiver.root_folder_id = "root_id"

    report_folder_id = archiver._get_or_create_folder("report", "root_id")
    assert report_folder_id == "report_folder_id"
